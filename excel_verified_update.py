"""
Verified Excel-to-rekordbox updater.

This is the reusable version of the one-off updater that was used to update the
library from `rekordbox_energy_rating_colors_updated.xlsx`.

It updates:
- Rating
- Color
- MyTag Genre from Genre_Normalized
- MyTag Situation from Energy_Label / Set Role
- MyTag Components from Elements + Mood

It does NOT update the metadata Genre field.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import random
import shutil
import sys
import uuid
from pathlib import Path
from typing import Any

import openpyxl

from pyrekordbox.db6 import (
    DjmdArtist,
    DjmdColor,
    DjmdContent,
    DjmdMyTag,
    DjmdSongMyTag,
    Rekordbox6Database,
)

DEFAULT_DB = Path(os.environ.get("REKORDBOX_DB_PATH", r"C:\Users\Admin\AppData\Roaming\Pioneer\rekordbox\master.db"))
DEFAULT_SHEET = "Tracks_Energy"
BACKUP_ROOT = Path("backups")

HEADER_ALIASES = {
    "no": ["#", "No"],
    "artist": ["Artist", "Исполнитель"],
    "title": ["Title", "Track", "Название дорожки"],
    "rating": ["Rating", "Рейтинг"],
    "color": ["Color"],
    "genre_tag": ["Genre_Normalized", "Genre Tag"],
    "role": ["Energy_Label", "Set Role", "Role"],
    "elements": ["Elements"],
    "mood": ["Mood"],
}

TEST_COMMENTS = {
    "Indie Dance; WARM; Instrumental; Indie; Groovy",
    "Breaks; JOURNEY; Instrumental; Breaks; Atmospheric; Emotional; Journey",
    "Melodic House; JOURNEY; Instrumental; Atmospheric",
}


def utc_now() -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc)


def split_values(value: Any) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).replace(";", ",").split(",") if part.strip()]


def rating_from_cell(value: Any) -> int:
    text = str(value or "").strip()
    if "*" in text:
        return max(0, min(5, text.count("*")))
    return max(0, min(5, int(float(text)))) if text else 0


def header_index(headers: dict[str, int], key: str) -> int:
    for alias in HEADER_ALIASES[key]:
        if alias in headers:
            return headers[alias]
    raise RuntimeError(f"Missing required Excel column for {key}: {HEADER_ALIASES[key]}")


def optional_header_index(headers: dict[str, int], key: str) -> int | None:
    for alias in HEADER_ALIASES[key]:
        if alias in headers:
            return headers[alias]
    return None


def load_rows(xlsx_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers_row = next(ws.iter_rows(values_only=True))
    headers = {str(name).strip(): i for i, name in enumerate(headers_row) if name is not None}

    idx_no = optional_header_index(headers, "no")
    idx_artist = optional_header_index(headers, "artist")
    idx_title = header_index(headers, "title")
    idx_rating = header_index(headers, "rating")
    idx_color = optional_header_index(headers, "color")
    idx_genre = optional_header_index(headers, "genre_tag")
    idx_role = optional_header_index(headers, "role")
    idx_elements = optional_header_index(headers, "elements")
    idx_mood = optional_header_index(headers, "mood")

    rows = []
    for row in ws.iter_rows(values_only=True):
        title = (row[idx_title] or "").strip() if row[idx_title] else ""
        if not title or title == headers_row[idx_title]:
            continue
        no = row[idx_no] if idx_no is not None else len(rows) + 1
        rows.append(
            {
                "no": int(no) if str(no).isdigit() else len(rows) + 1,
                "artist": row[idx_artist] if idx_artist is not None else None,
                "title": title,
                "rating": rating_from_cell(row[idx_rating]),
                "color": row[idx_color] if idx_color is not None else None,
                "tags": {
                    "Genre": split_values(row[idx_genre]) if idx_genre is not None else [],
                    "Situation": split_values(row[idx_role]) if idx_role is not None else [],
                    "Components": (split_values(row[idx_elements]) if idx_elements is not None else [])
                    + (split_values(row[idx_mood]) if idx_mood is not None else []),
                },
            }
        )
    return rows


def make_backup(db_path: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = BACKUP_ROOT / f"rekordbox_master_before_excel_update_{stamp}"
    backup_dir.mkdir(parents=True, exist_ok=False)
    for suffix in ["", "-wal", "-shm"]:
        src = Path(str(db_path) + suffix)
        if src.exists():
            shutil.copy2(src, backup_dir / src.name)
    return backup_dir


def max_local_usn(db: Rekordbox6Database) -> int:
    values = []
    for model in (DjmdContent, DjmdMyTag, DjmdSongMyTag):
        for (value,) in db.session.query(model.rb_local_usn).all():
            if value is not None:
                values.append(int(value))
    return max(values or [0]) + 1


def next_seq_by_parent(db: Rekordbox6Database) -> dict[str, int]:
    seqs = {}
    for tag in db.query(DjmdMyTag).all():
        seqs[tag.ParentID] = max(seqs.get(tag.ParentID, 0), tag.Seq or 0)
    return seqs


def next_numeric_id(existing_ids: set[str]) -> str:
    while True:
        value = str(random.randint(10_000_000, 4_200_000_000))
        if value not in existing_ids:
            existing_ids.add(value)
            return value


def new_mytag(parent_id: str, seq: int, name: str, local_usn: int, existing_ids: set[str]) -> DjmdMyTag:
    now = utc_now()
    return DjmdMyTag(
        ID=next_numeric_id(existing_ids),
        UUID=str(uuid.uuid4()),
        Seq=seq,
        Name=name,
        Attribute=0,
        ParentID=parent_id,
        rb_data_status=0,
        rb_local_data_status=0,
        rb_local_deleted=0,
        rb_local_synced=0,
        usn=None,
        rb_local_usn=local_usn,
        created_at=now,
        updated_at=now,
    )


def new_song_mytag(content_id: str, tag_id: str, local_usn: int) -> DjmdSongMyTag:
    now = utc_now()
    return DjmdSongMyTag(
        ID=str(uuid.uuid4()),
        UUID=str(uuid.uuid4()),
        MyTagID=tag_id,
        ContentID=content_id,
        TrackNo=None,
        rb_data_status=0,
        rb_local_data_status=0,
        rb_local_deleted=0,
        rb_local_synced=0,
        usn=None,
        rb_local_usn=local_usn,
        created_at=now,
        updated_at=now,
    )


def build_plan(db: Rekordbox6Database, rows: list[dict[str, Any]]) -> tuple[list[tuple[dict[str, Any], DjmdContent]], list[dict[str, Any]]]:
    artist_names = {artist.ID: artist.Name for artist in db.query(DjmdArtist).all()}
    contents = db.query(DjmdContent).all()
    content_by_key = {}
    for content in contents:
        key = ((content.Title or "").strip(), artist_names.get(content.ArtistID))
        content_by_key.setdefault(key, []).append(content)

    matched = []
    skipped = []
    for item in rows:
        candidates = content_by_key.get((item["title"], item["artist"]), [])
        if not candidates and item["artist"] is None:
            candidates = [content for content in contents if (content.Title or "").strip() == item["title"]]
        if len(candidates) != 1:
            skipped.append(
                {
                    "no": item["no"],
                    "artist": item["artist"],
                    "title": item["title"],
                    "matches": len(candidates),
                }
            )
            continue
        matched.append((item, candidates[0]))
    return matched, skipped


def apply_update(db_path: Path, rows: list[dict[str, Any]], dry_run: bool) -> dict[str, Any]:
    db = Rekordbox6Database(path=db_path)
    matched, skipped_matches = build_plan(db, rows)
    colors = {color.Commnt: color.ID for color in db.query(DjmdColor).all()}
    parents = {tag.Name: tag.ID for tag in db.query(DjmdMyTag).filter(DjmdMyTag.ParentID == "root").all()}
    required_parents = {"Genre", "Situation", "Components"}
    missing_parents = sorted(required_parents - set(parents))
    if missing_parents:
        raise RuntimeError(f"Missing MyTag parent folders: {missing_parents}")

    skipped_colors = []
    for item, _content in matched:
        if item["color"] and item["color"] not in colors:
            skipped_colors.append({"no": item["no"], "title": item["title"], "color": item["color"]})

    summary = {
        "mode": "dry-run" if dry_run else "apply",
        "rows": len(rows),
        "matched": len(matched),
        "skipped_matches": skipped_matches,
        "skipped_colors": skipped_colors,
    }
    if dry_run:
        return summary

    summary["backup_dir"] = str(make_backup(db_path))

    tag_by_parent_name = {}
    for tag in db.query(DjmdMyTag).filter(DjmdMyTag.rb_local_deleted == 0).all():
        tag_by_parent_name[(tag.ParentID, tag.Name)] = tag

    managed_parent_ids = {parents[name] for name in required_parents}
    managed_tag_ids = {
        tag.ID for tag in db.query(DjmdMyTag).filter(DjmdMyTag.ParentID.in_(managed_parent_ids)).all()
    }
    links_by_content = {}
    for link in db.query(DjmdSongMyTag).all():
        links_by_content.setdefault(link.ContentID, []).append(link)

    existing_ids = {tag.ID for tag in db.query(DjmdMyTag).all()}
    seqs = next_seq_by_parent(db)
    local_usn = max_local_usn(db)
    created_tags = []
    created_links = 0
    removed_links = 0
    updated_content = 0
    cleaned_comments = []

    for item, content in matched:
        content_changed = False
        if content.Rating != item["rating"]:
            content.Rating = item["rating"]
            content_changed = True
        if item["color"] in colors and content.ColorID != colors[item["color"]]:
            content.ColorID = colors[item["color"]]
            content_changed = True
        if (content.Commnt or "") in TEST_COMMENTS:
            content.Commnt = ""
            cleaned_comments.append({"no": item["no"], "title": content.Title})
            content_changed = True
        if content_changed:
            content.updated_at = utc_now()
            content.rb_local_usn = local_usn
            local_usn += 1
            updated_content += 1

        desired_tag_ids = set()
        for parent_name, names in item["tags"].items():
            parent_id = parents[parent_name]
            for name in names:
                tag = tag_by_parent_name.get((parent_id, name))
                if tag is None:
                    seqs[parent_id] = seqs.get(parent_id, 0) + 1
                    tag = new_mytag(parent_id, seqs[parent_id], name, local_usn, existing_ids)
                    local_usn += 1
                    db.add(tag)
                    tag_by_parent_name[(parent_id, name)] = tag
                    managed_tag_ids.add(tag.ID)
                    created_tags.append({"parent": parent_name, "name": name, "id": tag.ID})
                desired_tag_ids.add(tag.ID)

        existing_links = links_by_content.get(content.ID, [])
        existing_managed = {link.MyTagID: link for link in existing_links if link.MyTagID in managed_tag_ids}

        for tag_id, link in existing_managed.items():
            if tag_id not in desired_tag_ids:
                db.delete(link)
                removed_links += 1

        for tag_id in desired_tag_ids:
            if tag_id in existing_managed:
                continue
            link = new_song_mytag(content.ID, tag_id, local_usn)
            local_usn += 1
            db.add(link)
            links_by_content.setdefault(content.ID, []).append(link)
            created_links += 1

    db.commit()
    summary.update(
        {
            "updated_content": updated_content,
            "created_tags": created_tags,
            "created_links": created_links,
            "removed_links": removed_links,
            "cleaned_comments": cleaned_comments,
        }
    )
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Verified Excel table updater for rekordbox.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    rows = load_rows(args.xlsx, args.sheet)
    summary = apply_update(args.db, rows, dry_run=not args.apply)
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
